import re
from openpyxl import load_workbook
from sqlalchemy.orm import Session
from app.models import Project, Question, Checkpoint, Video


def import_checkpoints_xlsx(db: Session, file_path: str, project_id: int) -> dict:
    wb = load_workbook(file_path, read_only=True)

    stats = {"questions": 0, "checkpoints": 0, "skipped": 0}

    if "原题" in wb.sheetnames:
        ws = wb["原题"]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        for row in rows:
            if not row[0]:
                continue
            q_id = str(row[0]).strip()
            existing = db.query(Question).filter(Question.question_id == q_id).first()
            if existing:
                stats["skipped"] += 1
                continue
            q = Question(
                question_id=q_id,
                project_id=project_id,
                prompt=str(row[2]) if row[2] else "",
                language=str(row[3]) if row[3] else None,
                preprocess_note=str(row[4]) if row[4] else None,
            )
            db.add(q)
            stats["questions"] += 1
        db.flush()

    if "检查点拆解" in wb.sheetnames:
        ws = wb["检查点拆解"]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        for row in rows:
            if not row[0] or not row[1]:
                continue
            q_id = str(row[0]).strip()
            cp_id = str(row[1]).strip()

            existing = db.query(Checkpoint).filter(Checkpoint.checkpoint_id == cp_id).first()
            if existing:
                stats["skipped"] += 1
                continue

            question = db.query(Question).filter(Question.question_id == q_id).first()
            if not question:
                continue

            seq_match = re.search(r"CP(\d+)", cp_id)
            seq = int(seq_match.group(1)) if seq_match else None

            cp = Checkpoint(
                checkpoint_id=cp_id,
                question_id=question.id,
                seq=seq,
                text=str(row[2]) if row[2] else "",
                ability_id=str(row[3]) if row[3] else None,
                ability_name=str(row[4]) if row[4] else None,
                tag_id=str(row[5]) if row[5] else None,
                tag_name=str(row[6]) if row[6] else None,
                min_success_line=str(row[7]) if row[7] else None,
                evidence_period=str(row[8]) if row[8] else None,
                preprocess_note=str(row[9]) if len(row) > 9 and row[9] else None,
            )
            db.add(cp)
            stats["checkpoints"] += 1

    db.commit()
    wb.close()
    return stats


def import_video_list(db: Session, videos: list[dict], project_id: int) -> dict:
    stats = {"imported": 0, "skipped": 0}
    project = db.query(Project).filter(Project.id == project_id).first()

    for v in videos:
        video_id = v.get("video_id")
        question_id_str = v.get("question_id")
        if not video_id or not question_id_str:
            continue

        existing = db.query(Video).filter(Video.video_id == video_id).first()
        if existing:
            stats["skipped"] += 1
            continue

        question = db.query(Question).filter(Question.question_id == question_id_str).first()
        if not question:
            stats["skipped"] += 1
            continue

        video = Video(
            video_id=video_id,
            question_id=question.id,
            model_version=v.get("model_version") or (project.model_version if project else None),
            oss_url=v.get("oss_url", ""),
            duration_sec=v.get("duration_sec"),
        )
        db.add(video)
        stats["imported"] += 1

    db.commit()
    return stats
