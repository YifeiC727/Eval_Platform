import re
from datetime import datetime
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form
from sqlalchemy.orm import Session
from sqlalchemy import func
import shutil
import os
from app.database import get_db
from app.models import QuestionBank, Question, Checkpoint, Video

router = APIRouter(prefix="/api/banks", tags=["question-banks"])

UPLOAD_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), "data", "uploads")
os.makedirs(UPLOAD_DIR, exist_ok=True)


@router.get("/")
def list_banks(db: Session = Depends(get_db)):
    banks = db.query(QuestionBank).order_by(QuestionBank.updated_at.desc()).all()
    result = []
    for b in banks:
        q_count = db.query(func.count(Question.id)).filter(Question.bank_id == b.id).scalar()
        cp_count = db.query(func.count(Checkpoint.id)).join(Question).filter(Question.bank_id == b.id).scalar()
        result.append({
            "id": b.id,
            "name": b.name,
            "version": b.version,
            "description": b.description or "",
            "question_count": q_count,
            "checkpoint_count": cp_count,
            "created_at": b.created_at.isoformat() if b.created_at else None,
            "updated_at": b.updated_at.isoformat() if b.updated_at else None,
        })
    return result


@router.post("/")
def create_bank(data: dict, db: Session = Depends(get_db)):
    bank = QuestionBank(
        name=data.get("name", ""),
        description=data.get("description", ""),
    )
    db.add(bank)
    db.commit()
    db.refresh(bank)
    return {"id": bank.id, "name": bank.name, "version": bank.version}


@router.get("/{bank_id}")
def get_bank(bank_id: int, db: Session = Depends(get_db)):
    bank = db.query(QuestionBank).filter(QuestionBank.id == bank_id).first()
    if not bank:
        raise HTTPException(404, "bank not found")
    questions = db.query(Question).filter(Question.bank_id == bank_id).all()
    return {
        "id": bank.id,
        "name": bank.name,
        "version": bank.version,
        "description": bank.description,
        "question_count": len(questions),
        "questions": [
            {
                "id": q.id,
                "question_id": q.question_id,
                "prompt": q.prompt,
                "checkpoint_count": len(q.checkpoints),
            }
            for q in questions[:50]
        ],
    }


@router.delete("/{bank_id}")
def delete_bank(bank_id: int, db: Session = Depends(get_db)):
    bank = db.query(QuestionBank).filter(QuestionBank.id == bank_id).first()
    if not bank:
        raise HTTPException(404, "bank not found")
    # Delete checkpoints, questions
    questions = db.query(Question).filter(Question.bank_id == bank_id).all()
    for q in questions:
        db.query(Checkpoint).filter(Checkpoint.question_id == q.id).delete(synchronize_session=False)
    db.query(Question).filter(Question.bank_id == bank_id).delete(synchronize_session=False)
    db.delete(bank)
    db.commit()
    return {"status": "deleted"}


@router.post("/{bank_id}/import")
async def import_to_bank(
    bank_id: int,
    file: UploadFile = File(...),
    mode: str = Form("append"),
    db: Session = Depends(get_db),
):
    """
    导入xlsx到题库。mode: append(追加) 或 replace(覆盖全部)
    """
    bank = db.query(QuestionBank).filter(QuestionBank.id == bank_id).first()
    if not bank:
        raise HTTPException(404, "bank not found")

    file_path = os.path.join(UPLOAD_DIR, file.filename)
    with open(file_path, "wb") as f:
        shutil.copyfileobj(file.file, f)

    from openpyxl import load_workbook
    try:
        wb = load_workbook(file_path, read_only=True)
    except Exception as e:
        raise HTTPException(400, f"无法读取文件: {e}")

    stats = {"questions_added": 0, "questions_updated": 0, "checkpoints_added": 0, "checkpoints_updated": 0, "skipped": 0}

    if mode == "replace":
        # Delete all existing data in this bank
        for q in db.query(Question).filter(Question.bank_id == bank_id).all():
            db.query(Checkpoint).filter(Checkpoint.question_id == q.id).delete(synchronize_session=False)
        db.query(Question).filter(Question.bank_id == bank_id).delete(synchronize_session=False)
        db.flush()

    # Parse 原题 sheet
    url_map = {}
    pe_url_map = {}
    if "原题" in wb.sheetnames:
        ws = wb["原题"]
        header = [str(c.value or "").strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]
        url_col = header.index("视频URL") if "视频URL" in header else -1
        pe_url_col = header.index("PE视频URL") if "PE视频URL" in header else -1
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        for row in rows:
            if not row[0]:
                continue
            q_id = str(row[0]).strip()
            existing = db.query(Question).filter(Question.question_id == q_id, Question.bank_id == bank_id).first()
            if existing:
                existing.prompt = str(row[2]) if row[2] else existing.prompt
                existing.language = str(row[3]) if row[3] else existing.language
                existing.preprocess_note = str(row[4]) if row[4] else existing.preprocess_note
                stats["questions_updated"] += 1
            else:
                q = Question(
                    question_id=q_id,
                    bank_id=bank_id,
                    prompt=str(row[2]) if row[2] else "",
                    language=str(row[3]) if row[3] else None,
                    preprocess_note=str(row[4]) if row[4] else None,
                )
                db.add(q)
                stats["questions_added"] += 1
            if url_col >= 0 and len(row) > url_col and row[url_col]:
                url_map[q_id] = str(row[url_col]).strip()
                q_obj = db.query(Question).filter(Question.question_id == q_id, Question.bank_id == bank_id).first()
                if q_obj:
                    q_obj.video_url = str(row[url_col]).strip()
            if pe_url_col >= 0 and len(row) > pe_url_col and row[pe_url_col]:
                pe_url_map[q_id] = str(row[pe_url_col]).strip()
                q_obj = db.query(Question).filter(Question.question_id == q_id, Question.bank_id == bank_id).first()
                if q_obj:
                    q_obj.pe_video_url = str(row[pe_url_col]).strip()
        db.flush()

    # Parse 检查点拆解 sheet
    if "检查点拆解" in wb.sheetnames:
        ws = wb["检查点拆解"]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        for row in rows:
            if not row[0] or not row[1]:
                continue
            q_id = str(row[0]).strip()
            cp_id = str(row[1]).strip()

            question = db.query(Question).filter(Question.question_id == q_id, Question.bank_id == bank_id).first()
            if not question:
                stats["skipped"] += 1
                continue

            seq_match = re.search(r"CP(\d+)", cp_id)
            seq = int(seq_match.group(1)) if seq_match else None

            existing_cp = db.query(Checkpoint).filter(
                Checkpoint.checkpoint_id == cp_id, Checkpoint.question_id == question.id
            ).first()

            if existing_cp:
                existing_cp.text = str(row[2]) if row[2] else existing_cp.text
                existing_cp.ability_id = str(row[3]) if row[3] else existing_cp.ability_id
                existing_cp.ability_name = str(row[4]) if row[4] else existing_cp.ability_name
                existing_cp.tag_id = str(row[5]) if row[5] else existing_cp.tag_id
                existing_cp.tag_name = str(row[6]) if row[6] else existing_cp.tag_name
                existing_cp.min_success_line = str(row[7]) if row[7] else existing_cp.min_success_line
                existing_cp.evidence_period = str(row[8]) if len(row) > 8 and row[8] else existing_cp.evidence_period
                stats["checkpoints_updated"] += 1
            else:
                cp = Checkpoint(
                    checkpoint_id=cp_id,
                    question_id=question.id,
                    seq=seq,
                    text=str(row[2]) if row[2] else "",
                    ability_id=str(row[3]) if row[3] else None,
                    ability_name=str(row[4]) if row[4] else None,
                    tag_id=str(row[5]) if row[5] else None,
                    tag_name=str(row[6]) if row[6] else None,
                    min_success_line=str(row[7]) if row[7] else None,
                    evidence_period=str(row[8]) if len(row) > 8 and row[8] else None,
                    preprocess_note=str(row[9]) if len(row) > 9 and row[9] else None,
                )
                db.add(cp)
                stats["checkpoints_added"] += 1

    # Fallback: scan all sheets for "视频URL" column if url_map is still empty
    if not url_map:
        for sn in wb.sheetnames:
            ws = wb[sn]
            header = [str(c.value or "").strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]
            if "视频URL" in header and "题目ID" in header:
                url_col = header.index("视频URL")
                qid_col = header.index("题目ID")
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not row or not row[qid_col]:
                        continue
                    q_id = str(row[qid_col]).strip()
                    if q_id in url_map:
                        continue
                    if url_col < len(row) and row[url_col]:
                        url_map[q_id] = str(row[url_col]).strip()
                        q_obj = db.query(Question).filter(Question.question_id == q_id, Question.bank_id == bank_id).first()
                        if q_obj:
                            q_obj.video_url = str(row[url_col]).strip()
                break

    # Update bank version and timestamp
    bank.version += 1
    bank.updated_at = datetime.utcnow()

    # Sync video URLs to existing Video records that reference these questions
    if url_map:
        videos_updated = 0
        for q_id, url in url_map.items():
            question = db.query(Question).filter(Question.question_id == q_id, Question.bank_id == bank_id).first()
            if question:
                videos = db.query(Video).filter(Video.question_id == question.id, (Video.oss_url == None) | (Video.oss_url == "")).all()
                for v in videos:
                    v.oss_url = url
                    videos_updated += 1
        stats["videos_url_synced"] = videos_updated

    db.commit()
    wb.close()

    stats["new_version"] = bank.version
    stats["url_map_count"] = len(url_map)
    return {"status": "ok", "bank_id": bank_id, "url_map": url_map, **stats}
