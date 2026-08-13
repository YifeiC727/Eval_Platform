import io
from collections import Counter, defaultdict
from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from app.database import get_db
from app.models import FinalResult, Checkpoint, Question, Video, Assignment, Annotation, User

router = APIRouter(prefix="/api/export", tags=["export"])

SCORE_MAP = {"C": 1.0, "R": 0.3, "N": 0.0}
FAIL_CODE_NAMES = {
    "F01": "要求遗漏", "F02": "语义/指令错误", "F03": "数量/绑定错误",
    "F04": "结构/解剖错误", "F05": "动作动态错误", "F06": "交互/接触错误",
    "F07": "物理/因果错误", "F08": "时序错误", "F09": "一致性错误",
    "F10": "镜头/构图错误", "F11": "视觉呈现错误",
    "F010": "镜头/构图错误", "F011": "视觉呈现错误",
}


def _style_header(ws, row=1):
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="E8ECF0", end_color="E8ECF0", fill_type="solid")
    for cell in ws[row]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")


@router.get("/my-annotations")
def export_my_annotations(user_id: int = None, db: Session = Depends(get_db)):
    """标注员导出自己的标注结果"""
    if not user_id:
        return {"error": "user_id required"}

    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        return {"error": "user not found"}

    wb = Workbook()
    ws = wb.active
    ws.title = "我的标注"
    ws.append(["题目ID", "检查点ID", "检查点内容", "最低成功线", "能力ID", "能力名称", "我的判定", "备注", "视频URL"])
    _style_header(ws)

    assignments = db.query(Assignment).filter(
        Assignment.annotator_id == user_id,
        Assignment.status.in_(["submitted", "completed"]),
    ).all()

    for assignment in assignments:
        video = assignment.video
        question = video.question if video else None
        anns = db.query(Annotation).filter(Annotation.assignment_id == assignment.id).all()
        ann_map = {a.checkpoint_id: a for a in anns}

        checkpoints = db.query(Checkpoint).filter(
            Checkpoint.question_id == question.id
        ).order_by(Checkpoint.seq).all() if question else []

        for cp in checkpoints:
            ann = ann_map.get(cp.id)
            ws.append([
                question.question_id if question else "",
                cp.checkpoint_id,
                cp.text,
                cp.min_success_line or "",
                cp.ability_id or "",
                cp.ability_name or "",
                ann.score if ann else "",
                ann.note if ann else "",
                video.oss_url if video else "",
            ])

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"{user.username}_annotations.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/results")
def export_results(project_id: int = None, batch_id: int = None, db: Session = Depends(get_db)):
    from app.models import EvalBatch

    # Check if this is a PE batch
    is_pe = False
    if batch_id:
        batch = db.query(EvalBatch).filter(EvalBatch.id == batch_id).first()
        if batch and batch.eval_mode == "pe":
            is_pe = True

    if is_pe:
        return _export_pe_results(batch_id, db)
    else:
        return _export_base_results(project_id, batch_id, db)


def _export_pe_results(batch_id: int, db: Session):
    """PE评测专用导出：A/B得分对比 + 增益Δ + 胜平负 + 原因分布"""
    from app.models import EvalBatch
    wb = Workbook()

    batch = db.query(EvalBatch).filter(EvalBatch.id == batch_id).first()

    # ========== Sheet 1: PE总览指标 ==========
    ws1 = wb.active
    ws1.title = "PE总览"
    ws1.append(["指标", "值", "说明"])
    _style_header(ws1)

    # Collect per-question A/B scores
    videos = db.query(Video).filter(Video.batch_id == batch_id).all()
    question_scores = []

    for video in videos:
        q = video.question
        if not q:
            continue

        # Get finalized annotations for this video
        assignments = db.query(Assignment).filter(
            Assignment.video_id == video.id,
            Assignment.status == "submitted",
        ).all()

        # Collect A/B scores from annotations
        a_scores = []
        b_scores = []
        pe_comparison = None
        pe_reason = None
        pe_gsb_raw = None

        for asgn in assignments:
            if asgn.pe_comparison:
                pe_gsb_raw = asgn.pe_comparison
                # pe_comparison is JSON: {"dynamics": "b_better", ..., "overall": "b_better"}
                try:
                    import json as _json
                    gsb = _json.loads(asgn.pe_comparison)
                    pe_comparison = gsb.get("overall", "tie")
                except (ValueError, TypeError):
                    pe_comparison = asgn.pe_comparison  # fallback: old single-value format
            if asgn.pe_reason:
                try:
                    import json as _json
                    reasons = _json.loads(asgn.pe_reason)
                    # Get the overall reason or first available
                    pe_reason = reasons.get("overall") or next(iter(reasons.values()), None) if isinstance(reasons, dict) else asgn.pe_reason
                except (ValueError, TypeError):
                    pe_reason = asgn.pe_reason

            anns = db.query(Annotation).filter(Annotation.assignment_id == asgn.id).all()
            for ann in anns:
                if ann.score in SCORE_MAP:
                    if ann.target == "A":
                        a_scores.append(SCORE_MAP[ann.score])
                    elif ann.target == "B":
                        b_scores.append(SCORE_MAP[ann.score])

        if a_scores or b_scores:
            a_avg = sum(a_scores) / len(a_scores) * 100 if a_scores else None
            b_avg = sum(b_scores) / len(b_scores) * 100 if b_scores else None
            delta = (b_avg - a_avg) if (a_avg is not None and b_avg is not None) else None
            question_scores.append({
                "qid": q.question_id,
                "prompt": q.prompt,
                "a_score": a_avg,
                "b_score": b_avg,
                "delta": delta,
                "comparison": pe_comparison,
                "reason": pe_reason,
                "gsb_raw": pe_gsb_raw,
                "a_n": len(a_scores),
                "b_n": len(b_scores),
            })

    # Calculate summary metrics
    valid_qs = [q for q in question_scores if q["a_score"] is not None and q["b_score"] is not None]
    total_qs = len(valid_qs)

    if total_qs > 0:
        avg_a = sum(q["a_score"] for q in valid_qs) / total_qs
        avg_b = sum(q["b_score"] for q in valid_qs) / total_qs
        avg_delta = avg_b - avg_a

        b_better = sum(1 for q in valid_qs if q["comparison"] in ("B_better", "b_better"))
        tie = sum(1 for q in valid_qs if q["comparison"] in ("tie", "same_good", "same_bad"))
        b_worse = sum(1 for q in valid_qs if q["comparison"] in ("B_worse", "b_worse", "a_better", "A_better"))
        # Fallback: if no pe_comparison, use delta
        if b_better + tie + b_worse == 0:
            b_better = sum(1 for q in valid_qs if q["delta"] and q["delta"] > 5)
            tie = sum(1 for q in valid_qs if q["delta"] and abs(q["delta"]) <= 5)
            b_worse = sum(1 for q in valid_qs if q["delta"] and q["delta"] < -5)
    else:
        avg_a = avg_b = avg_delta = 0
        b_better = tie = b_worse = 0

    ws1.append(["A绝对得分", round(avg_a, 1), "A(直出)的检查点均分"])
    ws1.append(["B绝对得分", round(avg_b, 1), "B(PE)的检查点均分"])
    ws1.append(["PE增益Δ", round(avg_delta, 1), "B得分 - A得分"])
    ws1.append(["有效题目数", total_qs, "同时有A/B得分的题目"])
    ws1.append([""])
    ws1.append(["B更好(胜)", b_better, f"{round(b_better/total_qs*100,1) if total_qs else 0}%"])
    ws1.append(["基本持平(平)", tie, f"{round(tie/total_qs*100,1) if total_qs else 0}%"])
    ws1.append(["B更差(负)", b_worse, f"{round(b_worse/total_qs*100,1) if total_qs else 0}%"])

    # ========== Sheet 2: 按能力的PE增益 ==========
    ws2 = wb.create_sheet("能力维度PE增益")
    ws2.append(["能力ID", "能力名称", "A得分", "B得分", "增益Δ", "A有效n", "B有效n"])
    _style_header(ws2)

    # Collect per-ability A/B scores
    ability_a = defaultdict(lambda: {"name": "", "scores": []})
    ability_b = defaultdict(lambda: {"name": "", "scores": []})

    for video in videos:
        q = video.question
        if not q:
            continue
        assignments = db.query(Assignment).filter(
            Assignment.video_id == video.id, Assignment.status == "submitted"
        ).all()
        for asgn in assignments:
            anns = db.query(Annotation).filter(Annotation.assignment_id == asgn.id).all()
            for ann in anns:
                if ann.score not in SCORE_MAP:
                    continue
                cp = ann.checkpoint
                if not cp:
                    continue
                aid = cp.ability_id or "UNKNOWN"
                if ann.target == "A":
                    ability_a[aid]["name"] = cp.ability_name or ""
                    ability_a[aid]["scores"].append(SCORE_MAP[ann.score])
                elif ann.target == "B":
                    ability_b[aid]["name"] = cp.ability_name or ""
                    ability_b[aid]["scores"].append(SCORE_MAP[ann.score])

    all_aids = sorted(set(list(ability_a.keys()) + list(ability_b.keys())))
    ability_rows = []
    for aid in all_aids:
        a_data = ability_a[aid]
        b_data = ability_b[aid]
        a_n = len(a_data["scores"])
        b_n = len(b_data["scores"])
        a_score = round(sum(a_data["scores"]) / a_n * 100, 1) if a_n else None
        b_score = round(sum(b_data["scores"]) / b_n * 100, 1) if b_n else None
        delta = round(b_score - a_score, 1) if (a_score is not None and b_score is not None) else None
        name = a_data["name"] or b_data["name"]
        ability_rows.append((aid, name, a_score, b_score, delta, a_n, b_n))

    ability_rows.sort(key=lambda x: x[4] if x[4] is not None else 0)
    for row in ability_rows:
        ws2.append(list(row))

    # ========== Sheet 3: 题目明细（含GSB维度） ==========
    ws3 = wb.create_sheet("题目明细")

    is_dual = batch.annotation_mode == "dual"

    # Build header
    header = ["题目ID", "Prompt", "PE Prompt", "A得分", "B得分", "增益Δ",
              "动态与物理", "动态原因", "声音效果", "声音原因",
              "镜头语言", "镜头原因", "视觉美学", "美学原因",
              "综合评价", "综合原因"]
    if is_dual:
        header += ["标注员B-动态", "标注员B-声音", "标注员B-镜头", "标注员B-美学", "标注员B-综合", "标注员B-原因"]
    header += ["A检查点数", "B检查点数"]
    ws3.append(header)
    _style_header(ws3)

    import json as _json
    gsb_labels = {"a_better": "A更好", "b_better": "B更好", "same_good": "一样好", "same_bad": "一样差"}

    for video in videos:
        q = video.question
        if not q:
            continue

        # Find this question's score data
        q_data = next((qs for qs in question_scores if qs["qid"] == q.question_id), None)
        if not q_data:
            continue

        # Get GSB from all submitted assignments
        assignments = db.query(Assignment).filter(
            Assignment.video_id == video.id, Assignment.status == "submitted"
        ).all()

        gsb_list = []  # list of (gsb_dict, reasons_dict) per annotator
        for asgn in assignments:
            gsb = {}
            reasons = {}
            if asgn.pe_comparison:
                try:
                    gsb = _json.loads(asgn.pe_comparison)
                except:
                    gsb = {}
            if asgn.pe_reason:
                try:
                    reasons = _json.loads(asgn.pe_reason) if isinstance(asgn.pe_reason, str) else {}
                except:
                    reasons = {}
            if gsb:
                gsb_list.append((gsb, reasons))

        # First annotator's GSB
        gsb1 = gsb_list[0][0] if gsb_list else {}
        reasons1 = gsb_list[0][1] if gsb_list else {}

        row = [
            q_data["qid"],
            q_data["prompt"][:200],
            video.pe_prompt[:200] if video.pe_prompt else "",
            round(q_data["a_score"], 1) if q_data["a_score"] is not None else "",
            round(q_data["b_score"], 1) if q_data["b_score"] is not None else "",
            round(q_data["delta"], 1) if q_data["delta"] is not None else "",
            gsb_labels.get(gsb1.get("dynamics", ""), ""),
            reasons1.get("dynamics", ""),
            gsb_labels.get(gsb1.get("audio", ""), ""),
            reasons1.get("audio", ""),
            gsb_labels.get(gsb1.get("camera", ""), ""),
            reasons1.get("camera", ""),
            gsb_labels.get(gsb1.get("aesthetics", ""), ""),
            reasons1.get("aesthetics", ""),
            gsb_labels.get(gsb1.get("overall", ""), ""),
            reasons1.get("overall", ""),
        ]

        if is_dual:
            # Second annotator's GSB
            gsb2 = gsb_list[1][0] if len(gsb_list) > 1 else {}
            reasons2 = gsb_list[1][1] if len(gsb_list) > 1 else {}
            reason_parts = [f"{k}:{v}" for k, v in reasons2.items()] if reasons2 else []
            row += [
                gsb_labels.get(gsb2.get("dynamics", ""), ""),
                gsb_labels.get(gsb2.get("audio", ""), ""),
                gsb_labels.get(gsb2.get("camera", ""), ""),
                gsb_labels.get(gsb2.get("aesthetics", ""), ""),
                gsb_labels.get(gsb2.get("overall", ""), ""),
                "; ".join(reason_parts) if reason_parts else "",
            ]

        row += [q_data["a_n"], q_data["b_n"]]
        ws3.append(row)

    # ========== Sheet 4: 原始标注 ==========
    ws4 = wb.create_sheet("原始标注")
    ws4.append(["题目ID", "检查点ID", "能力ID", "能力名称", "标注员", "目标(A/B)", "判定", "失败码", "备注"])
    _style_header(ws4)

    ann_query = db.query(Annotation).join(Assignment, Annotation.assignment_id == Assignment.id).join(
        Video, Assignment.video_id == Video.id
    ).filter(Video.batch_id == batch_id)

    for ann in ann_query.all():
        asgn = ann.assignment
        video = asgn.video
        cp = ann.checkpoint
        q = cp.question if cp else None
        user = asgn.annotator
        ws4.append([
            q.question_id if q else "",
            cp.checkpoint_id if cp else "",
            cp.ability_id if cp else "",
            cp.ability_name if cp else "",
            user.display_name or user.username if user else "",
            ann.target or "",
            ann.score,
            ann.fail_code or "",
            ann.note or "",
        ])

    # ========== Sheet 5: 标注员统计 ==========
    ws5 = wb.create_sheet("标注员统计")
    ws5.append(["标注员", "姓名", "总任务", "已提交", "完成率%", "A标注数", "B标注数", "整体判定填写率%"])
    _style_header(ws5)

    annotators = db.query(User).filter(User.role.contains("annotator")).all()
    for user in annotators:
        assignments = db.query(Assignment).join(Video).filter(
            Assignment.annotator_id == user.id, Video.batch_id == batch_id
        ).all()
        if not assignments:
            continue
        total = len(assignments)
        submitted = sum(1 for a in assignments if a.status == "submitted")
        has_comparison = sum(1 for a in assignments if a.pe_comparison)

        assign_ids = [a.id for a in assignments]
        anns = db.query(Annotation).filter(Annotation.assignment_id.in_(assign_ids)).all()
        a_count = sum(1 for a in anns if a.target == "A")
        b_count = sum(1 for a in anns if a.target == "B")

        ws5.append([
            user.username, user.display_name or "",
            total, submitted,
            round(submitted / total * 100, 1) if total else 0,
            a_count, b_count,
            round(has_comparison / total * 100, 1) if total else 0,
        ])

    # Auto-width
    for ws in wb.worksheets:
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    batch_name = batch.name if batch else "PE"
    # Sanitize filename for HTTP header (avoid non-ASCII)
    safe_name = "".join(c if c.isascii() and c.isalnum() or c in "-_" else "_" for c in batch_name)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=PE_evaluation_report_{safe_name}.xlsx"},
    )


def _export_base_results(project_id: int, batch_id: int, db: Session):
    """基础评测导出（原有逻辑）"""
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "能力得分排名"
    ws1.append(["能力ID", "能力名称", "得分", "C数", "R数", "N数", "有效n", "C率%", "R率%", "N率%", "覆盖状态", "主要失败码", "主要三级标签"])
    _style_header(ws1)

    # Compute ability scores
    query = db.query(Checkpoint, FinalResult).join(
        FinalResult, FinalResult.checkpoint_id == Checkpoint.id
    ).filter(FinalResult.final_score.in_(["C", "R", "N"]))
    if batch_id:
        query = query.join(Video, FinalResult.video_id == Video.id).filter(Video.batch_id == batch_id)
    elif project_id:
        query = query.join(Question, Checkpoint.question_id == Question.id).filter(
            (Question.project_id == project_id) | (Question.bank_id == project_id)
        )

    ability_data = defaultdict(lambda: {"name": "", "scores": [], "c": 0, "r": 0, "n": 0, "fail_codes": [], "tags": []})
    for cp, fr in query.all():
        aid = cp.ability_id or "UNKNOWN"
        ability_data[aid]["name"] = cp.ability_name or ""
        ability_data[aid]["scores"].append(SCORE_MAP.get(fr.final_score, 0))
        if fr.final_score == "C":
            ability_data[aid]["c"] += 1
        elif fr.final_score == "R":
            ability_data[aid]["r"] += 1
        else:
            ability_data[aid]["n"] += 1
        if fr.final_fail_code:
            ability_data[aid]["fail_codes"].append(fr.final_fail_code)
        if cp.tag_id:
            ability_data[aid]["tags"].append(cp.tag_name or cp.tag_id)

    ability_rows = []
    for aid in sorted(ability_data.keys()):
        d = ability_data[aid]
        n = len(d["scores"])
        score = round(sum(d["scores"]) / n * 100, 1) if n > 0 else 0
        coverage = "正式排名" if n >= 10 else ("初步趋势" if n >= 5 else "证据不足")
        top_fc = Counter(d["fail_codes"]).most_common(1)
        top_tag = Counter(d["tags"]).most_common(1)
        ability_rows.append((
            aid, d["name"], score, d["c"], d["r"], d["n"], n,
            round(d["c"] / n * 100, 1) if n else 0,
            round(d["r"] / n * 100, 1) if n else 0,
            round(d["n"] / n * 100, 1) if n else 0,
            coverage,
            f'{top_fc[0][0]} {FAIL_CODE_NAMES.get(top_fc[0][0], "")}({top_fc[0][1]})' if top_fc else "",
            f'{top_tag[0][0]}({top_tag[0][1]})' if top_tag else "",
        ))

    ability_rows.sort(key=lambda x: x[2])
    for row in ability_rows:
        ws1.append(list(row))

    # ========== Sheet 2: 失败码分布 ==========
    ws2 = wb.create_sheet("失败码分布")
    ws2.append(["失败码", "名称", "次数", "占比%"])
    _style_header(ws2)

    all_fail_codes = []
    for d in ability_data.values():
        all_fail_codes.extend(d["fail_codes"])
    fc_counter = Counter(all_fail_codes)
    fc_total = sum(fc_counter.values())
    for code in sorted(fc_counter.keys()):
        ws2.append([code, FAIL_CODE_NAMES.get(code, ""), fc_counter[code],
                    round(fc_counter[code] / fc_total * 100, 1) if fc_total else 0])

    # ========== Sheet 3: 三级标签诊断 ==========
    ws3 = wb.create_sheet("三级标签诊断")
    ws3.append(["标签ID", "标签名称", "能力ID", "得分", "C数", "R数", "N数", "有效n"])
    _style_header(ws3)

    tag_query = db.query(Checkpoint, FinalResult).join(
        FinalResult, FinalResult.checkpoint_id == Checkpoint.id
    ).filter(FinalResult.final_score.in_(["C", "R", "N"]))
    if batch_id:
        tag_query = tag_query.join(Video, FinalResult.video_id == Video.id).filter(Video.batch_id == batch_id)
    elif project_id:
        tag_query = tag_query.join(Question, Checkpoint.question_id == Question.id).filter(
            (Question.project_id == project_id) | (Question.bank_id == project_id)
        )

    tag_data = defaultdict(lambda: {"name": "", "ability": "", "scores": [], "c": 0, "r": 0, "n": 0})
    for cp, fr in tag_query.all():
        tid = cp.tag_id or "UNKNOWN"
        tag_data[tid]["name"] = cp.tag_name or ""
        tag_data[tid]["ability"] = cp.ability_id or ""
        tag_data[tid]["scores"].append(SCORE_MAP.get(fr.final_score, 0))
        if fr.final_score == "C":
            tag_data[tid]["c"] += 1
        elif fr.final_score == "R":
            tag_data[tid]["r"] += 1
        else:
            tag_data[tid]["n"] += 1

    tag_rows = []
    for tid in sorted(tag_data.keys()):
        d = tag_data[tid]
        n = len(d["scores"])
        score = round(sum(d["scores"]) / n * 100, 1) if n else 0
        tag_rows.append((tid, d["name"], d["ability"], score, d["c"], d["r"], d["n"], n))
    tag_rows.sort(key=lambda x: x[3])
    for row in tag_rows:
        ws3.append(list(row))

    # ========== Sheet 4: 标注员统计 ==========
    ws4 = wb.create_sheet("标注员统计")
    ws4.append(["标注员", "姓名", "总任务", "已提交", "完成率%", "标注总数", "C数", "R数", "N数", "C率%", "R率%", "N率%"])
    _style_header(ws4)

    annotators = db.query(User).filter(User.role.contains("annotator")).all()
    for user in annotators:
        assign_query = db.query(Assignment).filter(Assignment.annotator_id == user.id)
        if batch_id:
            assign_query = assign_query.join(Video, Assignment.video_id == Video.id).filter(Video.batch_id == batch_id)

        assignments = assign_query.all()
        total_tasks = len(assignments)
        if total_tasks == 0:
            continue
        submitted = sum(1 for a in assignments if a.status == "submitted")

        # Get annotations for this user's assignments in scope
        assign_ids = [a.id for a in assignments]
        anns = db.query(Annotation).filter(Annotation.assignment_id.in_(assign_ids)).all() if assign_ids else []
        c_count = sum(1 for a in anns if a.score == "C")
        r_count = sum(1 for a in anns if a.score == "R")
        n_count = sum(1 for a in anns if a.score == "N")
        total_anns = c_count + r_count + n_count

        ws4.append([
            user.username, user.display_name or "",
            total_tasks, submitted,
            round(submitted / total_tasks * 100, 1) if total_tasks else 0,
            total_anns, c_count, r_count, n_count,
            round(c_count / total_anns * 100, 1) if total_anns else 0,
            round(r_count / total_anns * 100, 1) if total_anns else 0,
            round(n_count / total_anns * 100, 1) if total_anns else 0,
        ])

    # ========== Sheet 5: 题目明细 ==========
    ws5 = wb.create_sheet("题目明细")
    ws5.append(["题目ID", "Prompt", "检查点数", "题目得分", "完成率%", "C数", "R数", "N数", "主要失败能力"])
    _style_header(ws5)

    questions = db.query(Question)
    if batch_id:
        q_ids = [v.question_id for v in db.query(Video).filter(Video.batch_id == batch_id).all()]
        questions = questions.filter(Question.id.in_(q_ids))
    elif project_id:
        questions = questions.filter((Question.project_id == project_id) | (Question.bank_id == project_id))

    for q in questions.all():
        cps = q.checkpoints
        q_scores = []
        q_c = q_r = q_n = 0
        fail_abilities = []
        for cp in cps:
            fr_query = db.query(FinalResult).filter(FinalResult.checkpoint_id == cp.id)
            if batch_id:
                fr_query = fr_query.join(Video, FinalResult.video_id == Video.id).filter(Video.batch_id == batch_id)
            fr = fr_query.first()
            if fr and fr.final_score in ("C", "R", "N"):
                q_scores.append(SCORE_MAP[fr.final_score])
                if fr.final_score == "C":
                    q_c += 1
                elif fr.final_score == "R":
                    q_r += 1
                else:
                    q_n += 1
                if fr.final_score in ("R", "N"):
                    fail_abilities.append(cp.ability_id or "")

        k = len(q_scores)
        total_score = sum(q_scores)
        completion = round(total_score / k * 100, 1) if k else 0
        top_fail_ability = Counter(fail_abilities).most_common(1)

        ws5.append([
            q.question_id, q.prompt[:200],
            len(cps), round(total_score, 1), completion,
            q_c, q_r, q_n,
            top_fail_ability[0][0] if top_fail_ability else "",
        ])

    # ========== Sheet 6: 最终定案明细 ==========
    ws6 = wb.create_sheet("定案明细")
    ws6.append(["视频ID", "题目ID", "检查点ID", "能力ID", "能力名称", "三级标签", "最终判定", "失败码", "定案方式", "分值"])
    _style_header(ws6)

    fr_query = db.query(FinalResult).join(Checkpoint, FinalResult.checkpoint_id == Checkpoint.id)
    if batch_id:
        fr_query = fr_query.join(Video, FinalResult.video_id == Video.id).filter(Video.batch_id == batch_id)
    elif project_id:
        fr_query = fr_query.join(Question, Checkpoint.question_id == Question.id).filter(
            (Question.project_id == project_id) | (Question.bank_id == project_id)
        )

    for fr in fr_query.all():
        cp = fr.checkpoint
        q = cp.question
        v = fr.video
        ws6.append([
            v.video_id if v else "",
            q.question_id if q else "",
            cp.checkpoint_id,
            cp.ability_id or "",
            cp.ability_name or "",
            cp.tag_name or "",
            fr.final_score,
            fr.final_fail_code or "",
            fr.method,
            SCORE_MAP.get(fr.final_score, 0),
        ])

    # ========== Sheet 7: 原始标注记录 ==========
    ws7 = wb.create_sheet("原始标注")
    ws7.append(["视频ID", "题目ID", "检查点ID", "标注员", "角色", "判定", "失败码", "备注"])
    _style_header(ws7)

    ann_query = db.query(Annotation).join(Assignment, Annotation.assignment_id == Assignment.id)
    if batch_id:
        ann_query = ann_query.join(Video, Assignment.video_id == Video.id).filter(Video.batch_id == batch_id)
    elif project_id:
        ann_query = ann_query.join(Video, Assignment.video_id == Video.id).join(
            Question, Video.question_id == Question.id
        ).filter((Question.project_id == project_id) | (Question.bank_id == project_id))

    for ann in ann_query.all():
        asgn = ann.assignment
        video = asgn.video
        cp = ann.checkpoint
        q = cp.question if cp else None
        user = asgn.annotator
        ws7.append([
            video.video_id if video else "",
            q.question_id if q else "",
            cp.checkpoint_id if cp else "",
            user.display_name or user.username if user else "",
            asgn.role,
            ann.score,
            ann.fail_code or "",
            ann.note or "",
        ])

    # Auto-width for all sheets
    for ws in wb.worksheets:
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=V6_evaluation_report.xlsx"},
    )
